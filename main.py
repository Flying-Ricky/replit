from fastapi import FastAPI, File, UploadFile
from fastapi.responses import FileResponse
import tempfile, zipfile, os
import pandas as pd
from openpyxl import load_workbook
import re

app = FastAPI(title="Invoice Checker API")

def extract_vat_eori(path):
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    text = str(ws["B5"].value) if ws["B5"].value else ""
    vat = re.search(r"\bGB\d{9}\b", text)
    eori = re.search(r"\bGB\d{12}\b", text)
    return (vat.group(0) if vat else "", eori.group(0) if eori else "")

@app.post("/process_batch")
async def process_batch(file: UploadFile = File(...)):
    tmpdir = tempfile.mkdtemp()
    zip_path = os.path.join(tmpdir, file.filename)
    with open(zip_path, "wb") as f:
        f.write(await file.read())

    with zipfile.ZipFile(zip_path, 'r') as z:
        z.extractall(tmpdir)

    results = []
    for root, dirs, files in os.walk(tmpdir):
        for f in files:
            if f.lower().endswith((".xls", ".xlsx")):
                path = os.path.join(root, f)
                vat, eori = extract_vat_eori(path)
                results.append({"File": f, "VAT": vat, "EORI": eori})

    df = pd.DataFrame(results)
    out_path = os.path.join(tmpdir, "batch_results.xlsx")
    df.to_excel(out_path, index=False)

    return FileResponse(out_path, filename="batch_results.xlsx")
